from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.shortcuts import render

class PortalLoginView(LoginView):
    template_name = 'portal/login.html'

    # After a successful login, go to the dashboard
    def get_success_url(self):
        return self.request.GET.get('next', '/portal/')

@login_required
def portal_dashboard(request):
    return render(request, 'portal/dashboard.html')
@login_required
def leave_detail(request):
    return render(request, 'portal/leave_detail.html')

@login_required
def payroll_detail(request):
    return render(request, 'portal/payroll_detail.html')

@login_required
def reviews_detail(request):
    return render(request, 'portal/reviews_detail.html')
from django.http import FileResponse
from reportlab.pdfgen import canvas
from io import BytesIO

@login_required
def generate_payslip_pdf(request, month):
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 800, f"Payslip for {month}")
    p.drawString(100, 780, f"Employee: {request.user.username}")
    p.drawString(100, 760, "Amount: ₹50,000")  # Example data
    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'payslip_{month}.pdf')
from django.contrib.auth.views import LoginView

class PortalLoginView(LoginView):
    template_name = 'portal/login.html'

    def get_success_url(self):
        return self.request.GET.get('next', '/portal/dashboard/')
from django.contrib.auth.views import LoginView

class PortalLoginView(LoginView):
    template_name = 'portal/login.html'

    def get_success_url(self):
        return self.request.GET.get('next', '/portal/dashboard/')
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import LeaveRequest, Payroll, PerformanceReview
from .forms import LeaveRequestForm
from reportlab.pdfgen import canvas
from django.http import HttpResponse
# portal/views.py
from django.shortcuts import render, redirect
from .models import LeaveRequest
from .forms import LeaveForm  # Ensure you created this form
@login_required
def leave_detail(request):
    if request.method == 'POST':
        form = LeaveForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.user = request.user
            leave.full_clean()  # Validate model constraints
            leave.save()
            return redirect('leave-detail')
    else:
        form = LeaveForm()
    
    leaves = LeaveRequest.objects.filter(user=request.user)
    return render(request, 'portal/leave_detail.html', {
        'form': form,
        'leave_list': leaves
    })

@login_required
def payroll_detail(request):
    payslips = Payroll.objects.filter(user=request.user)
    return render(request, 'portal/payroll_detail.html', {'payslips': payslips})

@login_required
def generate_payslip_pdf(request, month):
    from io import BytesIO
    payslip = Payroll.objects.get(user=request.user, month=month)
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 800, f"Payslip for {month}")
    p.drawString(100, 780, f"Name: {request.user.get_full_name()}")
    p.drawString(100, 760, f"Basic Salary: {payslip.basic_salary}")
    p.drawString(100, 740, f"HRA: {payslip.hra}")
    p.drawString(100, 720, f"Deductions: {payslip.deductions}")
    p.drawString(100, 700, f"Net Pay: {payslip.net_pay}")
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def reviews_detail(request):
    reviews = PerformanceReview.objects.filter(user=request.user)
    return render(request, 'portal/reviews_detail.html', {'reviews': reviews})
from django.shortcuts import redirect, get_object_or_404
from .models import LeaveRequest
from django.contrib.auth.decorators import login_required

@login_required
def cancel_leave(request, leave_id):
    leave = get_object_or_404(LeaveRequest, id=leave_id, user=request.user)
    if leave.status == 'Pending':
        leave.status = 'Cancelled'
        leave.save()
    return redirect('leave-detail')
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Payroll

@login_required
def payroll_detail(request):
    selected_month = request.GET.get('month')

    # Get salary for selected month
    salary = None
    if selected_month:
        try:
            salary = Payroll.objects.get(user=request.user, month=selected_month)
        except Payroll.DoesNotExist:
            salary = None

    # Get all months with payslips
    payslip_months = Payroll.objects.filter(user=request.user).order_by('-month').values_list('month', flat=True)

    return render(request, 'portal/payroll_detail.html', {
        'salary': salary,
        'payslip_months': payslip_months,
        'selected_month': selected_month
    })

    
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import Font


# views.py
@login_required
def download_payslip_excel(request, month):
    payslip = Payroll.objects.get(user=request.user, month=month)
    wb = Workbook()
    ws = wb.active
    ws.title = "Payslip"

    ws['A1'] = f"Payslip – {month} 2025"
    ws['A1'].font = Font(size=14, bold=True)

    rows = [
        ("Company", "Zentrixon Technologies Pvt. Ltd."),
        ("Employee", request.user.get_full_name()),
        ("Employee ID", f"ZT{request.user.id}"),
        ("Designation", "Software Engineer"),
        ("Dept", "IT"),
        ("Pay Period", f"01-{month}-2025 to 30-{month}-2025"),
        ("Bank A/C", "XXXX-XXXX-XXXX-5678"),
        ("Basic Pay", payslip.basic),
        ("HRA", payslip.hra),
        ("Allowances", payslip.allowance),
        ("Deductions", payslip.deductions),
        ("Net Pay", payslip.net),
    ]
    row = 3
    for label, val in rows:
        ws[f'A{row}'], ws[f'B{row}'] = label, val
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Payslip_{month}.xlsx'
    wb.save(response)
    return response

@login_required
def generate_payslip_pdf(request, month):
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas

    payslip = Payroll.objects.get(user=request.user, month=month)

    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 800, f"Payslip for {month}")
    p.drawString(100, 780, f"Name: {request.user.get_full_name()}")
    p.drawString(100, 760, f"Basic Pay: ₹{payslip.basic}")
    p.drawString(100, 740, f"HRA: ₹{payslip.hra}")
    p.drawString(100, 720, f"Allowances: ₹{payslip.allowance}")
    p.drawString(100, 700, f"Deductions: ₹{payslip.deductions}")
    p.drawString(100, 680, f"Net Pay: ₹{payslip.net}")
    p.showPage()
    p.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@login_required
def download_payslip_excel(request, month):
    payslip = Payroll.objects.get(user=request.user, month=month)
    wb = Workbook()
    ws = wb.active
    ws.title = "Payslip"

    ws['A1'] = f"Payslip – {month} 2025"
    ws['A1'].font = Font(size=14, bold=True)

    rows = [
        ("Company", "Zentrixon Technologies Pvt. Ltd."),
        ("Employee", request.user.get_full_name()),
        ("Employee ID", f"ZT{request.user.id}"),
        ("Designation", "Software Engineer"),
        ("Dept", "IT"),
        ("Pay Period", f"01-{month}-2025 to 30-{month}-2025"),
        ("Bank A/C", "XXXX-XXXX-XXXX-5678"),
        ("Basic Pay", f"₹{payslip.basic:,.0f}"),
        ("HRA", f"₹{payslip.hra:,.0f}"),
        ("Allowances", f"₹{payslip.allowance:,.0f}"),
        ("Deductions", f"₹{payslip.deductions:,.0f}"),
        ("Net Pay", f"₹{payslip.net:,.0f}"),
    ]
    row = 3
    for label, val in rows:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = val
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Payslip_{month}.xlsx'
    wb.save(response)
    return response

from django.contrib import messages

@login_required
def leave_detail(request):
    if request.method == 'POST':
        form = LeaveForm(request.POST)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.user = request.user
            leave.save()
            messages.success(request, "Leave request submitted successfully!")
            return redirect('leave-detail')
    else:
        form = LeaveForm()

    leaves = LeaveRequest.objects.filter(user=request.user)
    return render(request, 'portal/leave_detail.html', {
        'form': form,
        'leave_list': leaves
    })

@login_required
def reviews_detail(request):
    reviews = PerformanceReview.objects.filter(user=request.user).order_by('-date')
    ratings = [review.rating for review in reviews]
    months = [review.date.strftime("%b %Y") for review in reviews]

    total_rating = sum(ratings) / len(ratings) if ratings else 0
    latest_review = reviews.first() if reviews else None

    return render(request, 'portal/reviews_detail.html', {
        'reviews': reviews,
        'ratings': ratings,
        'months': months,
        'total_rating': total_rating,
        'latest_review': latest_review
    })
